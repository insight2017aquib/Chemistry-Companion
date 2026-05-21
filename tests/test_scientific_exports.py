"""Normalized scientific workbook export tests."""

from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.app import app
from exports.excel.workbook_builder import build_workbook_bytes
from exports.exporters import CsvExporter, JsonExporter
from exports.schemas.batch_export_schema import build_batch_export_payload
from exports.schemas.workbook_models import ALL_WORKBOOK_SHEETS


def _records():
    return [
        {
            "success": True,
            "input": {"smiles": "c1ccccc1", "name": "Benzene"},
            "molecule": {
                "name": "Benzene",
                "smiles": "c1ccccc1",
                "formula": "C6H6",
                "molecular_weight": 78.114,
                "exact_mass": 78.04695,
                "num_atoms": 12,
                "num_heavy_atoms": 6,
                "num_rings": 1,
                "is_aromatic": True,
            },
            "descriptors": {
                "formula": "C6H6",
                "molecular_weight": 78.114,
                "exact_mass": 78.04695,
                "logp": 1.69,
                "tpsa": 0.0,
                "hbd": 0,
                "hba": 0,
                "rotatable_bonds": 0,
                "ring_count": 1,
                "heavy_atom_count": 6,
                "functional_groups": {"aromatic_ring": 1},
            },
            "functional_group_report": {
                "matches": [
                    {
                        "key": "aromatic_ring",
                        "name": "Aromatic ring",
                        "category": "aromatic",
                        "count": 1,
                        "atom_indices": [[0, 1, 2, 3, 4, 5]],
                    }
                ]
            },
            "ir_prediction": {
                "bands": [
                    {"label": "Aromatic C-H stretch", "lower_cm1": 3000, "upper_cm1": 3100, "intensity": "medium"}
                ]
            },
            "proton_nmr_prediction": {
                "signals": [
                    {"label": "Aromatic H", "shift_ppm": 7.27, "multiplicity": "s", "integration": 6}
                ]
            },
            "carbon_nmr_prediction": {
                "environments": [
                    {"label": "Aromatic C", "ppm_range": [125.0, 130.0], "carbon_count": 6}
                ]
            },
        },
        {
            "success": False,
            "input": {"smiles": "BAD###", "name": "Bad row"},
            "error": "Could not parse SMILES",
        },
    ]


def _workbook(profile: str = "full"):
    content = build_workbook_bytes(build_batch_export_payload(_records()), profile=profile)
    return load_workbook(BytesIO(content))


def test_full_workbook_has_required_scientific_sheets():
    wb = _workbook("full")
    assert wb.sheetnames == ALL_WORKBOOK_SHEETS
    assert wb.properties.title == "Chemistry Companion Scientific Export"


def test_workbook_sheets_have_filters_freezes_and_header_style():
    wb = _workbook("full")
    ws = wb["Descriptors"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == ws.dimensions
    assert ws.sheet_properties.tabColor.rgb.endswith("805AD5")
    assert ws["A1"].fill.fgColor.rgb.endswith("1F4E78")
    assert ws["E2"].number_format == "0.000"


def test_partial_failures_are_preserved_without_blocking_success_rows():
    wb = _workbook("full")
    assert wb["Molecules"].max_row == 2
    assert wb["Failed_Entries"].max_row == 2
    assert wb["Failed_Entries"]["F2"].value == "Could not parse SMILES"


def test_csv_and_json_exporters_consume_normalized_payload():
    payload = build_batch_export_payload(_records())
    csv_text = CsvExporter().to_string(payload)
    json_text = JsonExporter().to_string(payload)

    assert "functional_group_report" not in csv_text
    assert "aromatic_ring" not in csv_text.splitlines()[0]
    assert '"molecules"' in json_text
    assert '"failures"' in json_text


def test_large_dataset_workbook_builds_summary_rows():
    records = [_records()[0] for _ in range(250)]
    wb = load_workbook(BytesIO(build_workbook_bytes(build_batch_export_payload(records), profile="minimal")))
    assert wb["Summary"].max_row == 251
    assert wb["Molecules"].max_row == 251


def test_export_preview_endpoint_returns_structured_response():
    with TestClient(app) as client:
        response = client.post(
            "/api/export/preview",
            json={"data": _records(), "format": "xlsx", "profile": "spectroscopy"},
        )
    assert response.status_code == 200
    data = response.json()
    assert data["summary"]["total_records"] == 2
    assert "Proton_NMR" in data["sheets"]


def test_export_download_endpoint_streams_workbook():
    with TestClient(app) as client:
        response = client.post(
            "/api/export/download",
            json={"data": _records(), "format": "xlsx", "profile": "full"},
        )
    assert response.status_code == 200
    assert "spreadsheet" in response.headers["content-type"]
    assert response.headers["x-export-records"] == "2"
    assert load_workbook(BytesIO(response.content)).sheetnames == ALL_WORKBOOK_SHEETS
