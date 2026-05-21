"""
scripts/run_tool_comparison.py
==============================
Benchmarking suite comparing Chemistry Companion, RDKit, and Open Babel.

Measures:
- Speed (molecules / second)
- Quantitative features (Descriptors, Functional Groups)
- Qualitative features (Batch Workflow, Exports, Docking Prep, GUI)

Generates:
- outputs/comparison/tool_comparison.xlsx
- outputs/comparison/tool_comparison.md
- outputs/comparison/publication_plots/*.png
"""

import time
import json
import logging
import sys
from pathlib import Path
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("ToolComparison")

# Path setup
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Imports for CC
from rdkit import Chem
from rdkit.Chem import Descriptors
from rdkit.Chem import Fragments
import openbabel.pybel as pb

from core.molecule_utils import load_molecule
from core.descriptor_utils import compute_descriptors
import importlib.util as _ilu

def _load_module(name: str, rel_path: str):
    spec = _ilu.spec_from_file_location(name, ROOT / rel_path)
    mod  = _ilu.module_from_spec(spec)
    import sys as _sys
    _sys.modules.setdefault(name, mod)
    spec.loader.exec_module(mod)
    return mod

_fgd_mod = _load_module("spectra.functional_group_detector", "spectra/functional_group_detector.py")
detect_functional_groups = _fgd_mod.detect_functional_groups
get_registry = _fgd_mod.get_registry

BG         = "FFFFFF"
CARD       = "FFFFFF"
TEXT       = "000000"
BORDER_CLR = "DEE2E6"
C_BLUE   = "1F77B4"
C_GREEN  = "2CA02C"
C_PURPLE = "9467BD"

plt.rcParams.update({
    "figure.facecolor":  f"#{BG}",
    "axes.facecolor":    f"#{CARD}",
    "axes.edgecolor":    "#000000",
    "axes.labelcolor":   "#000000",
    "text.color":        "#000000",
    "xtick.color":       "#000000",
    "ytick.color":       "#000000",
    "grid.color":        "#E0E0E0",
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         10,
    "legend.facecolor":  f"#{CARD}",
    "legend.edgecolor":  "#000000",
    "savefig.dpi":       300,
    "savefig.format":    "svg",
})

# --- Benchmarking Functions ---

def benchmark_cc(smiles_list, iterations):
    start = time.time()
    for _ in range(iterations):
        for s in smiles_list:
            try:
                mol_rec = load_molecule(s)
                if mol_rec.rdkit_mol:
                    desc = compute_descriptors(mol_rec.rdkit_mol)
                    fgs = detect_functional_groups(mol_rec.rdkit_mol)
            except Exception:
                pass
    end = time.time()
    return end - start

def benchmark_rdkit(smiles_list, iterations):
    start = time.time()
    for _ in range(iterations):
        for s in smiles_list:
            try:
                m = Chem.MolFromSmiles(s)
                if m:
                    # Calculate all descriptors
                    desc = Descriptors.CalcMolDescriptors(m)
                    # Count fragments (which act as basic functional groups)
                    for func_name in dir(Fragments):
                        if func_name.startswith("fr_"):
                            getattr(Fragments, func_name)(m)
            except Exception:
                pass
    end = time.time()
    return end - start

def benchmark_openbabel(smiles_list, iterations):
    start = time.time()
    for _ in range(iterations):
        for s in smiles_list:
            try:
                m = pb.readstring("smi", s)
                if m:
                    desc = m.calcdesc()
            except Exception:
                pass
    end = time.time()
    return end - start

def run_benchmarks():
    csv_path = ROOT / "data" / "benchmark_molecules.csv"
    if not csv_path.exists():
        logger.error("Benchmark CSV not found.")
        return None

    df = pd.read_csv(csv_path)
    smiles_list = df["smiles"].dropna().tolist()
    
    # Using 50 iterations for robust timing
    iterations = 50
    total_mols = len(smiles_list) * iterations
    
    logger.info(f"Running speed benchmarks on {total_mols} molecules ( {len(smiles_list)} * {iterations} )...")
    
    time_cc = benchmark_cc(smiles_list, iterations)
    logger.info(f"Chemistry Companion: {time_cc:.2f} s")
    
    time_rdkit = benchmark_rdkit(smiles_list, iterations)
    logger.info(f"RDKit: {time_rdkit:.2f} s")
    
    time_ob = benchmark_openbabel(smiles_list, iterations)
    logger.info(f"Open Babel: {time_ob:.2f} s")
    
    speeds = {
        "Chemistry Companion": total_mols / time_cc,
        "RDKit": total_mols / time_rdkit,
        "Open Babel": total_mols / time_ob
    }
    
    return speeds

# --- Qualitative and Quantitative Data ---

def get_comparison_data(speeds):
    # Determine quantitative counts
    # Descriptors: CC computes ~14 core descriptors. RDKit computes ~210. OpenBabel computes ~24.
    # Functional Groups: CC has 27 categories. RDKit has ~85 fragments. OB relies on custom SMARTS (0 native).
    cc_fg_count = len(get_registry())
    rdkit_fg_count = len([x for x in dir(Fragments) if x.startswith("fr_")])
    
    quantitative = {
        "Tool": ["Chemistry Companion", "RDKit", "Open Babel"],
        "Speed (mols/sec)": [speeds["Chemistry Companion"], speeds["RDKit"], speeds["Open Babel"]],
        "Descriptors (Native)": [14, len(Descriptors._descList), 24], 
        "Functional Group Classes": [cc_fg_count, rdkit_fg_count, 0] # OB has no native FG classes, just SMARTS
    }
    
    # Qualitative Matrix (0-3 scale: 0=None, 1=Basic, 2=Good, 3=Native/Comprehensive)
    qualitative = {
        "Feature": ["Batch Workflow", "Exports & Reports", "Docking Preparation", "GUI Integration"],
        "Chemistry Companion": [3, 3, 3, 3],
        "RDKit": [1, 1, 1, 1], # Needs custom loops, pandas exports, no PDBQT native, Jupyter only
        "Open Babel": [2, 1, 2, 2] # CLI batching, basic exports, native PDBQT, external Avogadro
    }
    
    return pd.DataFrame(quantitative), pd.DataFrame(qualitative)

# --- Plotting Functions ---

def _save_fig(fig, out_dir, name):
    path = out_dir / name
    fig.savefig(path, dpi=300, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    logger.info(f"Saved plot: {name}")
    return path

def plot_speed(df, out_dir):
    fig, ax = plt.subplots(figsize=(8, 5))
    colors = [f"#{C_BLUE}", f"#{C_GREEN}", f"#{C_PURPLE}"]
    bars = ax.bar(df["Tool"], df["Speed (mols/sec)"], color=colors, alpha=0.85)
    
    for bar in bars:
        h = bar.get_height()
        ax.annotate(f"{h:.0f} mols/s", xy=(bar.get_x() + bar.get_width()/2, h),
                    xytext=(0, 4), textcoords="offset points", ha="center", color=f"#{TEXT}")
                    
    ax.set_ylabel("Processing Speed (molecules / second)")
    ax.set_title("Performance Comparison\n(SMILES parsing, descriptors, functional groups)")
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    _save_fig(fig, out_dir, "01_speed_comparison.svg")

def plot_counts(df, out_dir):
    fig, ax = plt.subplots(figsize=(8, 5))
    x = np.arange(len(df["Tool"]))
    w = 0.35
    
    bars1 = ax.bar(x - w/2, df["Descriptors (Native)"], w, label="Descriptors", color=f"#{C_BLUE}", alpha=0.85)
    bars2 = ax.bar(x + w/2, df["Functional Group Classes"], w, label="Functional Groups", color=f"#{C_GREEN}", alpha=0.85)
    
    for bars in (bars1, bars2):
        for bar in bars:
            h = bar.get_height()
            ax.annotate(f"{h}", xy=(bar.get_x() + bar.get_width()/2, h),
                        xytext=(0, 4), textcoords="offset points", ha="center", color=f"#{TEXT}", fontsize=9)
                        
    ax.set_xticks(x)
    ax.set_xticklabels(df["Tool"])
    ax.set_ylabel("Count")
    ax.set_title("Native Features Comparison")
    ax.legend(framealpha=0.2)
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    _save_fig(fig, out_dir, "03_descriptor_fg_counts.svg")

def plot_radar(df_qual, out_dir):
    features = df_qual["Feature"].tolist()
    N = len(features)
    angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
    angles += angles[:1]
    
    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor(f"#{BG}")
    ax.set_facecolor(f"#{CARD}")
    
    ax.set_ylim(0, 3.5)
    ax.set_yticks([1, 2, 3])
    ax.set_yticklabels(["Basic", "Good", "Native"], color=f"#{TEXT}", alpha=0.7)
    
    colors = [f"#{C_BLUE}", f"#{C_GREEN}", f"#{C_PURPLE}"]
    tools = ["Chemistry Companion", "RDKit", "Open Babel"]
    
    for idx, tool in enumerate(tools):
        values = df_qual[tool].tolist()
        values += values[:1]
        ax.plot(angles, values, color=colors[idx], linewidth=2.5, label=tool)
        ax.fill(angles, values, color=colors[idx], alpha=0.15)
        
    ax.set_thetagrids(np.degrees(angles[:-1]), features, color=f"#{TEXT}", fontsize=10)
    ax.set_title("Qualitative Feature Radar (0=None, 3=Native/Comprehensive)", pad=20, color=f"#{TEXT}")
    ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1), framealpha=0.2)
    
    _save_fig(fig, out_dir, "02_feature_radar.svg")

# --- Exporters ---

def export_xlsx(df_quant, df_qual, out_dir):
    path = out_dir / "tool_comparison.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Benchmark"
    
    # Headers
    bold_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2A2D3E")
    center = Alignment(horizontal="center", vertical="center")
    
    ws.append(["Tool Comparison - Quantitative Metrics"])
    ws.append(df_quant.columns.tolist())
    for r in df_quant.values.tolist():
        # Round speed
        r[1] = round(r[1], 1)
        ws.append(r)
        
    ws.append([])
    ws.append(["Tool Comparison - Qualitative Matrix (0-3)"])
    ws.append(df_qual.columns.tolist())
    for r in df_qual.values.tolist():
        ws.append(r)
        
    for row in ws.iter_rows():
        for cell in row:
            if cell.row in [1, 2, 7, 8]:
                cell.font = bold_font
                cell.fill = fill
            cell.alignment = center
            
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
        
    wb.save(path)
    logger.info(f"Saved XLSX report: {path}")
    return path

def export_markdown(df_quant, df_qual, out_dir):
    path = out_dir / "tool_comparison.md"
    
    # Format dataframes
    df_q = df_quant.copy()
    df_q["Speed (mols/sec)"] = df_q["Speed (mols/sec)"].round(1)
    
    lines = [
        "# Tool Comparison Benchmark",
        "",
        "Comparing **Chemistry Companion**, **RDKit**, and **Open Babel**.",
        "",
        "## Quantitative Metrics",
        "",
        df_q.to_markdown(index=False),
        "",
        "## Qualitative Feature Matrix",
        "",
        "> Score scale: `0` (None/External), `1` (Basic/Manual), `2` (Good), `3` (Native/Comprehensive)",
        "",
        df_qual.to_markdown(index=False),
        "",
        "## Publication Plots",
        "",
        "![Speed Comparison](publication_plots/01_speed_comparison.svg)",
        "![Feature Radar](publication_plots/02_feature_radar.svg)",
        "![Descriptor & FG Counts](publication_plots/03_descriptor_fg_counts.svg)"
    ]
    
    path.write_text("\n".join(lines), encoding="utf-8")
    logger.info(f"Saved Markdown report: {path}")

# --- Main ---

def main():
    out_dir = ROOT / "outputs" / "comparison"
    plot_dir = out_dir / "publication_plots"
    plot_dir.mkdir(parents=True, exist_ok=True)
    
    speeds = run_benchmarks()
    if not speeds:
        return
        
    df_quant, df_qual = get_comparison_data(speeds)
    
    plot_speed(df_quant, plot_dir)
    plot_counts(df_quant, plot_dir)
    plot_radar(df_qual, plot_dir)
    
    export_xlsx(df_quant, df_qual, out_dir)
    export_markdown(df_quant, df_qual, out_dir)
    
    logger.info("Comparison benchmark complete.")

if __name__ == "__main__":
    main()
