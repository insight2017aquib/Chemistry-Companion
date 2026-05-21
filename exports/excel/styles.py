"""Openpyxl styles for publication-quality workbook exports."""

from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from exports.excel.formatters import number_format_for


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(color="1F2937", size=10)
ALT_FILL = PatternFill("solid", fgColor="F5F8FB")
FAIL_FILL = PatternFill("solid", fgColor="FDECEC")
THIN_SIDE = Side(style="thin", color="D7DEE8")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

TAB_COLORS = {
    "Summary": "1F4E78",
    "Molecules": "2F855A",
    "Descriptors": "805AD5",
    "Functional_Groups": "B7791F",
    "IR_Predictions": "C53030",
    "Proton_NMR": "2B6CB0",
    "Carbon_NMR": "285E61",
    "Failed_Entries": "C53030",
    "Metadata": "4A5568",
}


def apply_workbook_properties(wb: Workbook, profile_name: str) -> None:
    wb.properties.creator = "Chemistry Companion"
    wb.properties.title = "Chemistry Companion Scientific Export"
    wb.properties.subject = profile_name
    wb.properties.keywords = "chemistry, medicinal chemistry, descriptors, spectroscopy"
    wb.properties.comments = "Generated from normalized BatchExportPayload."
    wb.properties.created = datetime.now(timezone.utc).replace(tzinfo=None, microsecond=0)
    wb.properties.modified = wb.properties.created


def style_header_row(ws: Worksheet, headers: list[str], row_index: int = 1) -> None:
    for col_index in range(1, len(headers) + 1):
        cell = ws.cell(row=row_index, column=col_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = TABLE_BORDER
    ws.row_dimensions[row_index].height = 24


def style_data_rows(ws: Worksheet, keys: list[str], start_row: int = 2) -> None:
    for row_index in range(start_row, ws.max_row + 1):
        is_alt = (row_index - start_row) % 2 == 1
        for col_index, key in enumerate(keys, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = TABLE_BORDER
            if is_alt:
                cell.fill = ALT_FILL
            if ws.title == "Failed_Entries":
                cell.fill = FAIL_FILL
            number_format = number_format_for(key)
            if number_format and isinstance(cell.value, (int, float)):
                cell.number_format = number_format


def apply_table_features(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    if ws.max_column and ws.max_row:
        ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS.get(ws.title, "4A5568")


def autosize_columns(ws: Worksheet, max_width: int = 64) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), max_width)
