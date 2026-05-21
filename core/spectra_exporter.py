"""
core/spectra_exporter.py
=========================
Styled XLSX exporter and publication-quality plot generator for the
Spectra Validation Workflow.

Clearly separates HEURISTIC predictions from EXPERIMENTAL reference values
in every sheet, column header, plot legend, and annotation.

Exports
-------
  validation_summary.xlsx      — 7-sheet styled workbook
  publication_plots/           — 8 high-res PNG figures (300 dpi)
  validation_report.md         — structured Markdown report
"""

from __future__ import annotations

import json
import logging
import math
from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import seaborn as sns
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.spectra_validation import (
    SpectraValidationReport,
    SpectraDomainMetrics,
    DOMAIN_LABELS,
    DEFAULT_TOLERANCES,
)

logger = logging.getLogger(__name__)

# ── Design tokens ──────────────────────────────────────────────────────────────
BG         = "FFFFFF"
CARD       = "FFFFFF"
SUBCARD    = "F8F9FA"
TEXT       = "000000"
DIM        = "6C757D"
BORDER_CLR = "DEE2E6"

C_BLUE   = "1F77B4"
C_GREEN  = "2CA02C"
C_PURPLE = "9467BD"
C_RED    = "D62728"
C_ORANGE = "FF7F0E"
C_GOLD   = "E377C2"

C_GOOD   = "E6F4EA"
C_WARN   = "FFF3E0"
C_BAD    = "FCE8E6"

DOMAIN_COLORS = {"ir": C_BLUE, "proton": C_GREEN, "carbon": C_PURPLE}
DOMAIN_MUTED  = {"ir": "E8F4F8", "proton": "E9F5E9", "carbon": "F4E8F8"}

# Matplotlib publication style
plt.rcParams.update({
    "figure.facecolor":  f"#{BG}",
    "axes.facecolor":    f"#{CARD}",
    "axes.edgecolor":    "#000000",
    "axes.labelcolor":   "#000000",
    "text.color":        "#000000",
    "xtick.color":       "#000000",
    "ytick.color":       "#000000",
    "grid.color":        "#E0E0E0",
    "font.family":       "sans-serif",
    "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
    "font.size":         10,
    "axes.titlesize":    12,
    "axes.titleweight":  "bold",
    "legend.facecolor":  f"#{CARD}",
    "legend.edgecolor":  "#000000",
    "savefig.dpi":       300,
    "savefig.format":    "svg",
})

HEURISTIC_TAG    = "[HEURISTIC]"
EXPERIMENTAL_TAG = "[EXPERIMENTAL]"


# ══════════════════════════════════════════════════════════════════════════════
# XLSX helpers
# ══════════════════════════════════════════════════════════════════════════════

def _font(bold=False, color=TEXT, size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _border() -> Border:
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _hdr(cell, text, bg=CARD, fg=TEXT, bold=True, size=10):
    cell.value = text
    cell.font = _font(bold=bold, color=fg, size=size)
    cell.fill = _fill(bg)
    cell.border = _border()
    cell.alignment = _center()


def _cell(cell, value, bg=None, bold=False, color=TEXT, left=False):
    cell.value = value
    cell.font = _font(bold=bold, color=color)
    cell.border = _border()
    cell.alignment = _left() if left else _center()
    if bg:
        cell.fill = _fill(bg)


def _metric_bg(value: float | None, invert=False) -> str:
    if value is None:
        return CARD
    if invert:  # lower = better (for MAE/RMSE)
        return C_GOOD if value < 0.2 else (C_WARN if value < 0.5 else C_BAD)
    return C_GOOD if value >= 0.85 else (C_WARN if value >= 0.60 else C_BAD)


def _col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _save_fig(fig: plt.Figure, path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=300, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    logger.info("Saved plot -> %s", path)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Executive Summary
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_summary(wb: openpyxl.Workbook, report: SpectraValidationReport,
                   csv_path: str, tolerances: dict):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    _hdr(ws["A1"],
         "Spectra Validation Report  |  Heuristic Predictions vs Experimental Reference",
         bg="0B0E1A", fg=C_BLUE, size=13)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    _hdr(ws["A2"],
         f"Generated: {datetime.now():%Y-%m-%d %H:%M}   |   Source: {csv_path}   |   "
         f"Molecules: {report.total()}   |   NOTE: All predictions are HEURISTIC (rule-based, approximate)",
         bg="0F1117", fg=DIM, bold=False, size=9)

    # Tolerances
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    _hdr(ws[f"A{row}"], "Matching Tolerances Used", bg=SUBCARD, fg=C_GOLD)

    row += 1
    for col, (domain, tol) in enumerate(tolerances.items(), 1):
        _hdr(ws.cell(row, col), f"{DOMAIN_LABELS[domain]}\n±{tol}", bg=DOMAIN_MUTED[domain], fg=TEXT, size=10)

    # Metrics table header
    row += 2
    metrics_headers = [
        "Domain", "Avg MAE", "Avg RMSE",
        f"Coverage {HEURISTIC_TAG}", f"Coverage {EXPERIMENTAL_TAG}",
        "Total Matches", "Records OK", "Records Failed"
    ]
    for col, h in enumerate(metrics_headers, 1):
        _hdr(ws.cell(row, col), h, bg=CARD)

    row += 1
    for domain in ("ir", "proton", "carbon"):
        mae  = report.average_mae(domain)
        rmse = report.average_rmse(domain)
        covp = report.average_coverage_predicted(domain)
        cove = report.average_coverage_experimental(domain)
        bg = DOMAIN_MUTED[domain]
        vals = [
            (DOMAIN_LABELS[domain], bg, True),
            (f"{mae:.4f}"  if mae  is not None else "N/A", _metric_bg(mae,  invert=True), False),
            (f"{rmse:.4f}" if rmse is not None else "N/A", _metric_bg(rmse, invert=True), False),
            (f"{covp*100:.1f}%" if covp is not None else "N/A", _metric_bg(covp), False),
            (f"{cove*100:.1f}%" if cove is not None else "N/A", _metric_bg(cove), False),
            (str(report.total_matches(domain)), bg, False),
            (str(report.successes()), C_GOOD, False),
            (str(report.failures()), C_BAD if report.failures() > 0 else C_GOOD, False),
        ]
        for col, (val, cbg, left) in enumerate(vals, 1):
            _cell(ws.cell(row, col), val, bg=cbg, left=left)
        row += 1

    # Disclaimer
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    _hdr(ws[f"A{row}"],
         "DISCLAIMER: All spectral predictions are HEURISTIC (rule-based). "
         "Reference 'experimental' values are HEURISTIC-DERIVED (prediction + Gaussian noise). "
         "Not suitable for regulatory or publication use without independent verification.",
         bg="200A0A", fg="FF8080", bold=False, size=9)

    _col_widths(ws, [22, 12, 12, 20, 22, 14, 12, 14])
    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# Sheets 2-4 — Per-domain validation tables
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_domain(wb: openpyxl.Workbook, report: SpectraValidationReport,
                  domain: str, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    label = DOMAIN_LABELS[domain]
    color = DOMAIN_COLORS[domain]

    # Title
    ws.merge_cells("A1:M1")
    _hdr(ws["A1"], f"{label} Validation  |  {HEURISTIC_TAG} Predicted  vs  {EXPERIMENTAL_TAG} Reference",
         bg="0B0E1A", fg=color, size=12)
    ws.row_dimensions[1].height = 24

    headers = [
        "Compound", "SMILES",
        f"Predicted Peaks {HEURISTIC_TAG}", f"Experimental Peaks {EXPERIMENTAL_TAG}",
        "Matched", "MAE", "RMSE",
        f"Coverage {HEURISTIC_TAG}", f"Coverage {EXPERIMENTAL_TAG}",
        "Missing (FN)", "Extra (FP)", "Tolerance", "Status"
    ]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(2, col), h, bg=CARD)
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    for row_i, rec in enumerate(report.records, 3):
        metrics: SpectraDomainMetrics | None = getattr(rec, f"{domain}_metrics")
        bg = "161828" if row_i % 2 == 0 else "1A1D2E"

        name  = rec.molecule_name or ""
        smiles = rec.canonical_smiles or rec.input_smiles or ""
        status = "OK" if rec.success else f"FAILED: {rec.validation_error or ''}"
        status_bg = C_GOOD if rec.success else C_BAD
        status_fg = "80FF80" if rec.success else "FF8080"

        if metrics is None:
            row_data = [name, smiles, "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"]
        else:
            row_data = [
                name, smiles,
                str(metrics.predicted_count),
                str(metrics.experimental_count),
                str(metrics.matched_count),
                f"{metrics.mae:.4f}"  if metrics.mae  is not None else "N/A",
                f"{metrics.rmse:.4f}" if metrics.rmse is not None else "N/A",
                f"{metrics.coverage_predicted*100:.1f}%",
                f"{metrics.coverage_experimental*100:.1f}%",
                str(metrics.missing_experimental),
                str(metrics.extra_predicted),
                f"±{metrics.tolerance}",
            ]

        for col, val in enumerate(row_data, 1):
            cbg = bg
            if col == 6 and metrics and metrics.mae is not None:
                cbg = _metric_bg(metrics.mae, invert=True)
            elif col == 8 and metrics:
                cbg = _metric_bg(metrics.coverage_predicted)
            elif col == 9 and metrics:
                cbg = _metric_bg(metrics.coverage_experimental)
            _cell(ws.cell(row_i, col), val, bg=cbg, left=(col <= 2))

        # Status column
        _cell(ws.cell(row_i, 13), status, bg=status_bg, color=status_fg, bold=True)

    _col_widths(ws, [20, 30, 16, 18, 10, 10, 10, 18, 20, 12, 10, 10, 18])


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5 — Per-Molecule Full Detail
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_per_molecule(wb: openpyxl.Workbook, report: SpectraValidationReport):
    ws = wb.create_sheet("Per-Molecule Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:T1")
    _hdr(ws["A1"],
         f"Per-Molecule Summary  |  {HEURISTIC_TAG} vs {EXPERIMENTAL_TAG}  |  All Three Domains",
         bg="0B0E1A", fg=C_GOLD, size=12)
    ws.row_dimensions[1].height = 24

    domains = ("ir", "proton", "carbon")
    fixed_headers = ["Molecule", "SMILES", "Success"]
    domain_headers = []
    for d in domains:
        lbl = DOMAIN_LABELS[d]
        domain_headers += [f"{lbl} Pred", f"{lbl} Exp", f"{lbl} Match",
                           f"{lbl} MAE", f"{lbl} RMSE", f"{lbl} Cov%"]
    all_headers = fixed_headers + domain_headers

    for col, h in enumerate(all_headers, 1):
        _hdr(ws.cell(2, col), h, bg=CARD, size=9)
    ws.freeze_panes = "A3"

    for row_i, rec in enumerate(report.records, 3):
        bg = "161828" if row_i % 2 == 0 else "1A1D2E"
        suc = rec.success
        _cell(ws.cell(row_i, 1), rec.molecule_name or "", bg=bg, left=True)
        _cell(ws.cell(row_i, 2), rec.canonical_smiles or rec.input_smiles, bg=bg, left=True)
        _cell(ws.cell(row_i, 3), "OK" if suc else "FAIL",
              bg=C_GOOD if suc else C_BAD,
              color="80FF80" if suc else "FF8080", bold=True)
        col = 4
        for domain in domains:
            m: SpectraDomainMetrics | None = getattr(rec, f"{domain}_metrics")
            if m is None:
                for _ in range(6):
                    _cell(ws.cell(row_i, col), "-", bg=bg)
                    col += 1
            else:
                _cell(ws.cell(row_i, col),   m.predicted_count,   bg=bg)
                _cell(ws.cell(row_i, col+1), m.experimental_count, bg=bg)
                _cell(ws.cell(row_i, col+2), m.matched_count,       bg=bg)
                mae_bg = _metric_bg(m.mae, invert=True) if m.mae is not None else bg
                _cell(ws.cell(row_i, col+3),
                      f"{m.mae:.3f}" if m.mae is not None else "N/A", bg=mae_bg)
                _cell(ws.cell(row_i, col+4),
                      f"{m.rmse:.3f}" if m.rmse is not None else "N/A", bg=bg)
                cov_bg = _metric_bg(m.coverage_experimental)
                _cell(ws.cell(row_i, col+5),
                      f"{m.coverage_experimental*100:.1f}%", bg=cov_bg)
                col += 6

    _col_widths(ws, [20, 30, 8] + [10] * (len(all_headers) - 3))


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6 — Missing Predictions (FN)
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_missing(wb: openpyxl.Workbook, report: SpectraValidationReport):
    ws = wb.create_sheet("Missing Predictions")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    _hdr(ws["A1"],
         f"Missing Predictions (False Negatives)  |  {EXPERIMENTAL_TAG} peaks with no {HEURISTIC_TAG} match",
         bg="200A0A", fg="FF8080", size=11)

    headers = ["Molecule", "Domain", f"Experimental Value {EXPERIMENTAL_TAG}",
               f"Nearest Predicted {HEURISTIC_TAG}", "Gap", "Tolerance Used"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(2, col), h, bg=CARD)
    ws.freeze_panes = "A3"

    row_i = 3
    for rec in report.records:
        for domain in ("ir", "proton", "carbon"):
            m: SpectraDomainMetrics | None = getattr(rec, f"{domain}_metrics")
            if m is None or m.missing_experimental == 0:
                continue
            # Rebuild unmatched experimental peaks from details
            matched_exp = {d["experimental"] for d in m.details}
            # We don't have all experimental peaks stored, so report count + tolerance
            bg = "2A1010"
            _cell(ws.cell(row_i, 1), rec.molecule_name or rec.input_smiles, bg=bg, left=True)
            _cell(ws.cell(row_i, 2), DOMAIN_LABELS[domain], bg=bg)
            _cell(ws.cell(row_i, 3), f"{m.missing_experimental} peak(s) unmatched", bg=bg)
            _cell(ws.cell(row_i, 4), f"{m.extra_predicted} extra predicted", bg=bg)
            _cell(ws.cell(row_i, 5), "-", bg=bg)
            _cell(ws.cell(row_i, 6), f"±{m.tolerance}", bg=bg)
            row_i += 1

    if row_i == 3:
        ws.merge_cells("A3:F3")
        _cell(ws.cell(3, 1), "No missing predictions found — all experimental peaks matched.",
              bg=C_GOOD, color="80FF80", bold=True)

    _col_widths(ws, [22, 12, 28, 26, 10, 14])


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7 — Failures
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_failures(wb: openpyxl.Workbook, report: SpectraValidationReport):
    ws = wb.create_sheet("Failures")
    ws.sheet_view.showGridLines = False
    headers = ["Molecule", "Input SMILES", "Canonical SMILES", "Error"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(1, col), h, bg=C_BAD)
    failures = [r for r in report.records if not r.success]
    for row_i, rec in enumerate(failures, 2):
        bg = "2A1010"
        _cell(ws.cell(row_i, 1), rec.molecule_name or "",  bg=bg, left=True)
        _cell(ws.cell(row_i, 2), rec.input_smiles,          bg=bg, left=True)
        _cell(ws.cell(row_i, 3), rec.canonical_smiles or "", bg=bg, left=True)
        _cell(ws.cell(row_i, 4), rec.validation_error or "", bg=bg, left=True, color="FF8080")
    if not failures:
        _cell(ws.cell(2, 1), "No failures.", bg=C_GOOD, color="80FF80", bold=True)
    _col_widths(ws, [20, 35, 35, 55])


# ══════════════════════════════════════════════════════════════════════════════
# Master XLSX export
# ══════════════════════════════════════════════════════════════════════════════

def export_xlsx_styled(
    report: SpectraValidationReport,
    out_path: Path,
    csv_path: str = "",
    tolerances: dict | None = None,
) -> Path:
    """Build and save the 7-sheet styled XLSX validation report."""
    tolerances = tolerances or dict(DEFAULT_TOLERANCES)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, report, csv_path, tolerances)
    _sheet_domain(wb, report, "ir",     "IR Validation")
    _sheet_domain(wb, report, "proton", "1H NMR Validation")
    _sheet_domain(wb, report, "carbon", "13C NMR Validation")
    _sheet_per_molecule(wb, report)
    _sheet_missing(wb, report)
    _sheet_failures(wb, report)

    wb.save(out_path)
    logger.info("XLSX saved -> %s", out_path)
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# Publication plots
# ══════════════════════════════════════════════════════════════════════════════

def _collect_pairs(report: SpectraValidationReport, domain: str
                   ) -> tuple[list[float], list[float]]:
    """Return (predicted, experimental) value pairs from matched details."""
    pred_vals, exp_vals = [], []
    for rec in report.records:
        m: SpectraDomainMetrics | None = getattr(rec, f"{domain}_metrics")
        if m is None:
            continue
        for d in m.details:
            pred_vals.append(d["predicted"])
            exp_vals.append(d["experimental"])
    return pred_vals, exp_vals


def _collect_errors(report: SpectraValidationReport, domain: str) -> list[float]:
    errors = []
    for rec in report.records:
        m: SpectraDomainMetrics | None = getattr(rec, f"{domain}_metrics")
        if m:
            errors.extend(d["error"] for d in m.details)
    return errors


def plot_01_mae_rmse_overview(report: SpectraValidationReport, out_dir: Path) -> Path:
    domains = ["ir", "proton", "carbon"]
    labels  = [DOMAIN_LABELS[d] for d in domains]
    mae_vals  = [report.average_mae(d) or 0  for d in domains]
    rmse_vals = [report.average_rmse(d) or 0 for d in domains]

    x = np.arange(len(domains))
    w = 0.35
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor(f"#{BG}")

    b1 = ax.bar(x - w/2, mae_vals,  w, label=f"MAE {HEURISTIC_TAG}",
                color=[f"#{DOMAIN_COLORS[d]}" for d in domains], alpha=0.88)
    b2 = ax.bar(x + w/2, rmse_vals, w, label=f"RMSE {HEURISTIC_TAG}",
                color=[f"#{DOMAIN_COLORS[d]}" for d in domains], alpha=0.50, hatch="//")

    for bar in list(b1) + list(b2):
        h = bar.get_height()
        if h > 0:
            ax.annotate(f"{h:.3f}",
                        xy=(bar.get_x() + bar.get_width()/2, h),
                        xytext=(0, 4), textcoords="offset points",
                        ha="center", fontsize=8, color=f"#{TEXT}")

    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_ylabel("Error magnitude")
    ax.set_title(f"Average MAE & RMSE per Domain\n{HEURISTIC_TAG} predictions vs {EXPERIMENTAL_TAG} reference")
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    ax.legend(framealpha=0.2, fontsize=9)

    path = out_dir / "01_mae_rmse_overview.svg"
    _save_fig(fig, path)
    return path


def plot_02_coverage_heatmap(report: SpectraValidationReport, out_dir: Path) -> Path:
    domains = ["ir", "proton", "carbon"]
    names   = [r.molecule_name or r.input_smiles[:20] for r in report.records if r.success]
    data    = np.zeros((len(domains), len(names)))

    for di, domain in enumerate(domains):
        for ni, rec in enumerate([r for r in report.records if r.success]):
            m = getattr(rec, f"{domain}_metrics")
            data[di, ni] = m.coverage_experimental if m else 0

    fig, ax = plt.subplots(figsize=(max(12, len(names)*0.5), 4.5))
    fig.patch.set_facecolor(f"#{BG}")

    cmap = sns.color_palette("RdYlGn", as_cmap=True)
    sns.heatmap(
        data, ax=ax,
        xticklabels=names,
        yticklabels=[DOMAIN_LABELS[d] for d in domains],
        cmap=cmap, vmin=0, vmax=1,
        linewidths=0.5, linecolor=f"#{BORDER_CLR}",
        annot=True, fmt=".2f", annot_kws={"size": 7},
        cbar_kws={"label": f"Coverage {EXPERIMENTAL_TAG}"},
    )
    ax.set_title(f"Experimental Coverage Heatmap\n{EXPERIMENTAL_TAG} peaks matched by {HEURISTIC_TAG} predictions",
                 pad=12)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha="right", fontsize=7)
    plt.tight_layout()
    path = out_dir / "02_coverage_heatmap.svg"
    _save_fig(fig, path)
    return path


def _scatter_domain(report, domain, xlabel, ylabel, title, out_dir, filename) -> Path:
    pred, exp = _collect_pairs(report, domain)
    color = f"#{DOMAIN_COLORS[domain]}"

    fig, ax = plt.subplots(figsize=(7, 7))
    fig.patch.set_facecolor(f"#{BG}")

    ax.scatter(pred, exp, color=color, alpha=0.7, s=40, zorder=4,
               label=f"Matched peaks ({len(pred)})")

    if pred and exp:
        all_vals = pred + exp
        mn, mx = min(all_vals)*0.95, max(all_vals)*1.05
        ax.plot([mn, mx], [mn, mx], "--", color=f"#{C_GOLD}", linewidth=1.5,
                label="Identity line (perfect prediction)", zorder=5)

    ax.set_xlabel(f"{xlabel}  {HEURISTIC_TAG}")
    ax.set_ylabel(f"{ylabel}  {EXPERIMENTAL_TAG}")
    ax.set_title(title)
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.xaxis.grid(True, linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    ax.legend(framealpha=0.2, fontsize=9)

    # MAE annotation
    m = report.average_mae(domain)
    if m is not None:
        ax.annotate(f"Avg MAE = {m:.3f}", xy=(0.05, 0.92), xycoords="axes fraction",
                    fontsize=10, color=f"#{C_ORANGE}",
                    bbox=dict(boxstyle="round,pad=0.3", fc=f"#{CARD}", ec=f"#{BORDER_CLR}"))

    path = out_dir / filename
    _save_fig(fig, path)
    return path


def plot_03_ir_scatter(report, out_dir) -> Path:
    return _scatter_domain(
        report, "ir",
        "Predicted wavenumber (cm⁻¹)", "Experimental wavenumber (cm⁻¹)",
        f"IR: {HEURISTIC_TAG} vs {EXPERIMENTAL_TAG}\nPredicted mid-wavenumber vs Reference",
        out_dir, "03_ir_scatter.svg"
    )


def plot_04_1h_scatter(report, out_dir) -> Path:
    return _scatter_domain(
        report, "proton",
        "Predicted δ (ppm)", "Experimental δ (ppm)",
        f"1H NMR: {HEURISTIC_TAG} vs {EXPERIMENTAL_TAG}\nPredicted ppm vs Reference",
        out_dir, "04_1h_scatter.svg"
    )


def plot_05_13c_scatter(report, out_dir) -> Path:
    return _scatter_domain(
        report, "carbon",
        "Predicted δ (ppm)", "Experimental δ (ppm)",
        f"13C NMR: {HEURISTIC_TAG} vs {EXPERIMENTAL_TAG}\nPredicted ppm vs Reference",
        out_dir, "05_13c_scatter.svg"
    )


def plot_06_error_distributions(report, out_dir) -> Path:
    domains = ["ir", "proton", "carbon"]
    colors  = [f"#{DOMAIN_COLORS[d]}" for d in domains]

    fig, axes = plt.subplots(1, 3, figsize=(15, 5))
    fig.patch.set_facecolor(f"#{BG}")
    fig.suptitle(f"Absolute Error Distributions  {HEURISTIC_TAG}",
                 fontsize=13, color=f"#{TEXT}", y=1.02)

    for ax, domain, color in zip(axes, domains, colors):
        errors = _collect_errors(report, domain)
        if errors:
            ax.hist(errors, bins=20, color=color, alpha=0.7, edgecolor=f"#{BORDER_CLR}", zorder=3)
            try:
                from scipy.stats import gaussian_kde
                kde_x = np.linspace(0, max(errors), 200)
                kde = gaussian_kde(errors)
                ax2 = ax.twinx()
                ax2.plot(kde_x, kde(kde_x), color=color, linewidth=2.5, label="KDE")
                ax2.set_ylabel("Density", color=f"#{DIM}", fontsize=8)
                ax2.tick_params(axis="y", colors=f"#{DIM}", labelsize=7)
                ax2.set_facecolor(f"#{CARD}")
            except ImportError:
                pass
            mae = sum(errors) / len(errors)
            ax.axvline(mae, color=f"#{C_GOLD}", linestyle="--", linewidth=1.5,
                       label=f"Mean = {mae:.3f}")
            ax.legend(fontsize=8, framealpha=0.2)
        ax.set_title(DOMAIN_LABELS[domain])
        ax.set_xlabel("Absolute error")
        ax.set_ylabel("Count")
        ax.yaxis.grid(True, linestyle="--", alpha=0.3)
        ax.set_axisbelow(True)

    plt.tight_layout()
    path = out_dir / "06_error_distributions.svg"
    _save_fig(fig, path)
    return path


def plot_07_missing_predictions(report, out_dir) -> Path:
    names = []
    ir_fn, proton_fn, carbon_fn = [], [], []
    for rec in report.records:
        if not rec.success:
            continue
        names.append((rec.molecule_name or rec.input_smiles)[:20])
        ir_fn.append(rec.ir_metrics.missing_experimental if rec.ir_metrics else 0)
        proton_fn.append(rec.proton_metrics.missing_experimental if rec.proton_metrics else 0)
        carbon_fn.append(rec.carbon_metrics.missing_experimental if rec.carbon_metrics else 0)

    x = np.arange(len(names))
    w = 0.25
    fig, ax = plt.subplots(figsize=(max(12, len(names)*0.6), 5))
    fig.patch.set_facecolor(f"#{BG}")

    ax.bar(x - w, ir_fn,     w, label=f"IR {DOMAIN_LABELS['ir']}",
           color=f"#{DOMAIN_COLORS['ir']}", alpha=0.85)
    ax.bar(x,     proton_fn, w, label=f"1H NMR",
           color=f"#{DOMAIN_COLORS['proton']}", alpha=0.85)
    ax.bar(x + w, carbon_fn, w, label=f"13C NMR",
           color=f"#{DOMAIN_COLORS['carbon']}", alpha=0.85)

    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("Unmatched peaks (FN)")
    ax.set_title(f"Missing Predictions per Molecule\n{EXPERIMENTAL_TAG} peaks with no {HEURISTIC_TAG} match within tolerance")
    ax.yaxis.grid(True, linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    ax.legend(framealpha=0.2)
    plt.tight_layout()

    path = out_dir / "07_missing_predictions.svg"
    _save_fig(fig, path)
    return path


def plot_08_coverage_radar(report, out_dir) -> Path:
    domains = ["ir", "proton", "carbon"]
    labels  = [DOMAIN_LABELS[d] for d in domains]
    pred_c  = [report.average_coverage_predicted(d)    or 0 for d in domains]
    exp_c   = [report.average_coverage_experimental(d) or 0 for d in domains]

    N      = len(domains)
    angles = np.linspace(0, 2*np.pi, N, endpoint=False).tolist()
    angles_c = angles + [angles[0]]

    fig, ax = plt.subplots(figsize=(6.5, 6.5), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor(f"#{BG}")
    ax.set_facecolor(f"#{CARD}")

    ax.set_ylim(0, 1)
    ax.set_yticks([0.25, 0.5, 0.75, 1.0])
    ax.set_yticklabels(["25%", "50%", "75%", "100%"], color=f"#{DIM}", fontsize=8)
    ax.yaxis.grid(color=f"#{BORDER_CLR}", linestyle="--")
    ax.xaxis.grid(color=f"#{BORDER_CLR}", linestyle="--")

    for vals, color, label in [
        (pred_c, f"#{C_BLUE}",   f"Predicted coverage {HEURISTIC_TAG}"),
        (exp_c,  f"#{C_GREEN}",  f"Experimental coverage {EXPERIMENTAL_TAG}"),
    ]:
        vals_c = vals + [vals[0]]
        ax.plot(angles_c, vals_c, color=color, linewidth=2.5, zorder=4, label=label)
        ax.fill(angles_c, vals_c, color=color, alpha=0.15, zorder=3)

    ax.set_thetagrids(np.degrees(angles), labels, fontsize=11)
    ax.set_title(f"Coverage Radar\n{HEURISTIC_TAG} vs {EXPERIMENTAL_TAG}",
                 pad=22, fontsize=12, fontweight="bold")
    ax.legend(loc="upper right", bbox_to_anchor=(1.4, 1.15), framealpha=0.2, fontsize=9)

    path = out_dir / "08_coverage_radar.svg"
    _save_fig(fig, path)
    return path


def build_publication_plots(report: SpectraValidationReport, plot_dir: Path) -> dict[str, Path]:
    plt.rcParams["font.family"] = "Arial"
    plot_dir = Path(plot_dir)
    plot_dir.mkdir(parents=True, exist_ok=True)
    figures = {}
    for name, fn in [
        ("mae_rmse",       plot_01_mae_rmse_overview),
        ("coverage_hmap",  plot_02_coverage_heatmap),
        ("ir_scatter",     plot_03_ir_scatter),
        ("1h_scatter",     plot_04_1h_scatter),
        ("13c_scatter",    plot_05_13c_scatter),
        ("error_dist",     plot_06_error_distributions),
        ("missing",        plot_07_missing_predictions),
        ("radar",          plot_08_coverage_radar),
    ]:
        try:
            path = fn(report, plot_dir)
            figures[name] = path
            logger.info("Plot generated: %s", path.name)
        except Exception as exc:
            logger.warning("Plot '%s' failed: %s", name, exc)
    return figures


# ══════════════════════════════════════════════════════════════════════════════
# Markdown report
# ══════════════════════════════════════════════════════════════════════════════

def build_markdown_report_enhanced(
    report: SpectraValidationReport,
    tolerances: dict | None = None,
    plot_dir: Path | None = None,
) -> str:
    tolerances = tolerances or dict(DEFAULT_TOLERANCES)
    lines: list[str] = []

    lines += [
        "# Spectra Validation Report",
        "",
        "> **DISCLAIMER** — All spectral predictions in this report are **HEURISTIC**",
        "> (rule-based, functional-group derived, approximate). Reference 'experimental'",
        "> values are **HEURISTIC-DERIVED** (same predictions + Gaussian noise). They are",
        "> _NOT_ real experimental measurements. Do not use for regulatory or publication",
        "> purposes without independent verification.",
        "",
        f"**Generated:** {datetime.now():%Y-%m-%d %H:%M:%S}  ",
        f"**Molecules evaluated:** {report.total()}  ",
        f"**Successful:** {report.successes()}  |  **Failed:** {report.failures()}",
        "",
    ]

    # Tolerances
    lines += ["## Matching Tolerances", ""]
    lines += ["| Domain | Tolerance | Unit |", "| --- | --- | --- |"]
    units = {"ir": "cm⁻¹", "proton": "ppm", "carbon": "ppm"}
    for domain, tol in tolerances.items():
        lines.append(f"| {DOMAIN_LABELS[domain]} | ±{tol} | {units[domain]} |")
    lines.append("")

    # Per-domain summary
    lines += ["## Domain Metrics Summary", ""]
    lines += ["| Domain | Avg MAE | Avg RMSE | Pred Coverage | Exp Coverage | Total Matches |",
              "| --- | --- | --- | --- | --- | --- |"]
    for domain in ("ir", "proton", "carbon"):
        mae  = report.average_mae(domain)
        rmse = report.average_rmse(domain)
        cp   = report.average_coverage_predicted(domain)
        ce   = report.average_coverage_experimental(domain)
        tm   = report.total_matches(domain)
        lines.append(
            f"| {DOMAIN_LABELS[domain]} "
            f"| {mae:.4f} " if mae is not None else "| N/A "
            f"| {rmse:.4f} " if rmse is not None else "| N/A "
            f"| {cp*100:.1f}% " if cp is not None else "| N/A "
            f"| {ce*100:.1f}% " if ce is not None else "| N/A "
            f"| {tm} |"
        )
    lines.append("")

    # Per-domain detail sections
    for domain in ("ir", "proton", "carbon"):
        label = DOMAIN_LABELS[domain]
        unit  = units[domain]
        lines += [
            f"## {label} Validation",
            "",
            f"> **[HEURISTIC]** Predicted peaks vs **[EXPERIMENTAL]** reference ±{tolerances[domain]} {unit}",
            "",
            f"| Molecule | Pred Peaks [H] | Exp Peaks [E] | Matched | MAE | RMSE | Cov(pred) | Cov(exp) | Missing | Extra |",
            "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
        for rec in report.records:
            m: SpectraDomainMetrics | None = getattr(rec, f"{domain}_metrics")
            name = (rec.molecule_name or rec.input_smiles[:30]).replace("|", "\\|")
            if m is None:
                lines.append(f"| {name} | — | — | — | — | — | — | — | — | — |")
            else:
                lines.append(
                    f"| {name} "
                    f"| {m.predicted_count} "
                    f"| {m.experimental_count} "
                    f"| {m.matched_count} "
                    f"| {m.mae:.3f} " if m.mae is not None else "| N/A "
                    f"| {m.rmse:.3f} " if m.rmse is not None else "| N/A "
                    f"| {m.coverage_predicted*100:.1f}% "
                    f"| {m.coverage_experimental*100:.1f}% "
                    f"| {m.missing_experimental} "
                    f"| {m.extra_predicted} |"
                )
        lines.append("")

    # Failures
    if report.failures() > 0:
        lines += ["## Failures", "", "| Molecule | Error |", "| --- | --- |"]
        for rec in report.records:
            if not rec.success:
                name = (rec.molecule_name or rec.input_smiles[:30]).replace("|", "\\|")
                err  = (rec.validation_error or "").replace("|", "\\|")
                lines.append(f"| {name} | {err} |")
        lines.append("")

    # Figure links
    if plot_dir:
        lines += ["## Publication Plots", ""]
        fig_map = {
            "01_mae_rmse_overview.png":    "MAE & RMSE Overview",
            "02_coverage_heatmap.png":     "Coverage Heatmap",
            "03_ir_scatter.png":           "IR Scatter (Predicted vs Experimental)",
            "04_1h_scatter.png":           "1H NMR Scatter",
            "05_13c_scatter.png":          "13C NMR Scatter",
            "06_error_distributions.png":  "Error Distributions",
            "07_missing_predictions.png":  "Missing Predictions",
            "08_coverage_radar.png":       "Coverage Radar",
        }
        for fname, desc in fig_map.items():
            fpath = Path(plot_dir) / fname
            if fpath.exists():
                lines.append(f"![{desc}]({fpath})")
            lines.append("")

    return "\n".join(lines)
