"""
core/descriptor_benchmark.py
============================
Descriptor benchmarking utilities comparing Chemistry Companion descriptor outputs
against an RDKit reference implementation.
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from rdkit import Chem
from rdkit.Chem import Crippen, Descriptors, Lipinski, rdMolDescriptors

from core.descriptor_utils import DescriptorRecord, compute_descriptors
from core.molecule_utils import build_record, load_molecule


LOGP_TOLERANCE = 0.01
TPSA_TOLERANCE = 0.1


@dataclass(slots=True)
class DescriptorComparison:
    name: str | None
    smiles: str
    formula_cc: str | None
    formula_ref: str | None
    formula_match: bool
    mw_cc: float | None
    mw_ref: float | None
    mw_error: float | None
    logp_cc: float | None
    logp_ref: float | None
    logp_error: float | None
    logp_agreement: bool | None
    tpsa_cc: float | None
    tpsa_ref: float | None
    tpsa_error: float | None
    tpsa_agreement: bool | None
    rotatable_bonds_cc: int | None
    rotatable_bonds_ref: int | None
    rotatable_bonds_match: bool | None
    ring_count_cc: int | None
    ring_count_ref: int | None
    ring_count_match: bool | None


@dataclass(slots=True)
class DescriptorBenchmarkSummary:
    n_molecules: int
    formula_accuracy: float
    rotatable_bonds_accuracy: float
    ring_count_accuracy: float
    full_agreement: float
    mw_mae: float
    mw_rmse: float
    mw_max_error: float
    logp_mae: float
    logp_rmse: float
    logp_max_error: float
    logp_within_tolerance: float
    tpsa_mae: float
    tpsa_rmse: float
    tpsa_max_error: float
    tpsa_within_tolerance: float


def compute_rdkit_reference_descriptors(mol: Chem.Mol) -> DescriptorRecord:
    if mol is None:
        raise ValueError("Cannot compute RDKit reference descriptors for None molecule.")

    return DescriptorRecord(
        molecular_weight=float(Descriptors.MolWt(mol)),
        exact_mass=float(rdMolDescriptors.CalcExactMolWt(mol)),
        formula=str(rdMolDescriptors.CalcMolFormula(mol)),
        logp=float(Crippen.MolLogP(mol)),
        tpsa=float(rdMolDescriptors.CalcTPSA(mol)),
        hbd=int(Lipinski.NumHDonors(mol)),
        hba=int(Lipinski.NumHAcceptors(mol)),
        rotatable_bonds=int(Lipinski.NumRotatableBonds(mol)),
        ring_count=int(rdMolDescriptors.CalcNumRings(mol)),
        heavy_atom_count=int(mol.GetNumHeavyAtoms()),
        formal_charge=int(sum(atom.GetFormalCharge() for atom in mol.GetAtoms())),
        fraction_csp3=float(rdMolDescriptors.CalcFractionCSP3(mol)),
        functional_groups={},
        bertz_ct=None,
    )


def compare_descriptors(
    cc_desc: DescriptorRecord,
    rdkit_desc: DescriptorRecord,
    name: str | None = None,
    smiles: str | None = None,
    logp_tolerance: float = LOGP_TOLERANCE,
    tpsa_tolerance: float = TPSA_TOLERANCE,
) -> DescriptorComparison:
    if smiles is None:
        smiles = ""

    def diff(left: float | None, right: float | None) -> float | None:
        if left is None or right is None:
            return None
        return abs(left - right)

    formula_match = bool(cc_desc.formula == rdkit_desc.formula)
    mw_error = diff(cc_desc.molecular_weight, rdkit_desc.molecular_weight)
    logp_error = diff(cc_desc.logp, rdkit_desc.logp)
    tpsa_error = diff(cc_desc.tpsa, rdkit_desc.tpsa)

    rotatable_bonds_match = None
    if cc_desc.rotatable_bonds is not None and rdkit_desc.rotatable_bonds is not None:
        rotatable_bonds_match = cc_desc.rotatable_bonds == rdkit_desc.rotatable_bonds

    ring_count_match = None
    if cc_desc.ring_count is not None and rdkit_desc.ring_count is not None:
        ring_count_match = cc_desc.ring_count == rdkit_desc.ring_count

    return DescriptorComparison(
        name=name,
        smiles=smiles,
        formula_cc=cc_desc.formula,
        formula_ref=rdkit_desc.formula,
        formula_match=formula_match,
        mw_cc=cc_desc.molecular_weight,
        mw_ref=rdkit_desc.molecular_weight,
        mw_error=mw_error,
        logp_cc=cc_desc.logp,
        logp_ref=rdkit_desc.logp,
        logp_error=logp_error,
        logp_agreement=logp_error is not None and logp_error <= logp_tolerance,
        tpsa_cc=cc_desc.tpsa,
        tpsa_ref=rdkit_desc.tpsa,
        tpsa_error=tpsa_error,
        tpsa_agreement=tpsa_error is not None and tpsa_error <= tpsa_tolerance,
        rotatable_bonds_cc=cc_desc.rotatable_bonds,
        rotatable_bonds_ref=rdkit_desc.rotatable_bonds,
        rotatable_bonds_match=rotatable_bonds_match,
        ring_count_cc=cc_desc.ring_count,
        ring_count_ref=rdkit_desc.ring_count,
        ring_count_match=ring_count_match,
    )


def benchmark_molecule(
    smiles: str,
    name: str | None = None,
    logp_tolerance: float = LOGP_TOLERANCE,
    tpsa_tolerance: float = TPSA_TOLERANCE,
) -> DescriptorComparison:
    mol_rec = load_molecule(smiles=smiles)
    cc_desc = compute_descriptors(mol_rec.rdkit_mol, mw=mol_rec.mol_weight)
    rdkit_desc = compute_rdkit_reference_descriptors(mol_rec.rdkit_mol)
    return compare_descriptors(
        cc_desc=cc_desc,
        rdkit_desc=rdkit_desc,
        name=name or mol_rec.name,
        smiles=mol_rec.smiles,
        logp_tolerance=logp_tolerance,
        tpsa_tolerance=tpsa_tolerance,
    )


def benchmark_from_csv(
    csv_path: str | Path,
    smiles_column: str = "smiles",
    name_column: str = "name",
) -> list[DescriptorComparison]:
    import csv as _csv

    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(csv_path)

    comparisons: list[DescriptorComparison] = []
    with path.open(newline="", encoding="utf-8") as fh:
        reader = _csv.DictReader(fh)
        for row in reader:
            smiles = (row.get(smiles_column) or row.get(smiles_column.capitalize()) or "").strip()
            name = (row.get(name_column) or row.get(name_column.capitalize()) or None)
            if not smiles:
                continue
            comparisons.append(benchmark_molecule(smiles=smiles, name=name))
    return comparisons


def _statistics(values: list[float]) -> tuple[float, float, float]:
    if not values:
        return 0.0, 0.0, 0.0
    if len(values) == 1:
        return values[0], values[0], values[0]
    return mean(values), math.sqrt(sum(v * v for v in values) / len(values)), max(values)


def benchmark_summary(comparisons: list[DescriptorComparison]) -> DescriptorBenchmarkSummary:
    n = len(comparisons)
    if n == 0:
        raise ValueError("No comparisons provided for benchmark summary.")

    formula_matches = [1 for c in comparisons if c.formula_match]
    rotb_matches = [1 for c in comparisons if c.rotatable_bonds_match is True]
    rings_matches = [1 for c in comparisons if c.ring_count_match is True]
    full_matches = [
        1
        for c in comparisons
        if c.formula_match
        and c.rotatable_bonds_match is True
        and c.ring_count_match is True
        and c.logp_agreement is True
        and c.tpsa_agreement is True
    ]

    mw_errors = [c.mw_error for c in comparisons if c.mw_error is not None]
    logp_errors = [c.logp_error for c in comparisons if c.logp_error is not None]
    tpsa_errors = [c.tpsa_error for c in comparisons if c.tpsa_error is not None]

    return DescriptorBenchmarkSummary(
        n_molecules=n,
        formula_accuracy=100.0 * sum(formula_matches) / n,
        rotatable_bonds_accuracy=100.0 * sum(rotb_matches) / n,
        ring_count_accuracy=100.0 * sum(rings_matches) / n,
        full_agreement=100.0 * sum(full_matches) / n,
        mw_mae=mean(mw_errors) if mw_errors else 0.0,
        mw_rmse=math.sqrt(sum(e * e for e in mw_errors) / len(mw_errors)) if mw_errors else 0.0,
        mw_max_error=max(mw_errors) if mw_errors else 0.0,
        logp_mae=mean(logp_errors) if logp_errors else 0.0,
        logp_rmse=math.sqrt(sum(e * e for e in logp_errors) / len(logp_errors)) if logp_errors else 0.0,
        logp_max_error=max(logp_errors) if logp_errors else 0.0,
        logp_within_tolerance=100.0 * sum(1 for c in comparisons if c.logp_agreement is True) / n,
        tpsa_mae=mean(tpsa_errors) if tpsa_errors else 0.0,
        tpsa_rmse=math.sqrt(sum(e * e for e in tpsa_errors) / len(tpsa_errors)) if tpsa_errors else 0.0,
        tpsa_max_error=max(tpsa_errors) if tpsa_errors else 0.0,
        tpsa_within_tolerance=100.0 * sum(1 for c in comparisons if c.tpsa_agreement is True) / n,
    )


def comparisons_to_rows(comparisons: list[DescriptorComparison]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for comparison in comparisons:
        rows.append(
            {
                "name": comparison.name or "",
                "smiles": comparison.smiles,
                "formula_cc": comparison.formula_cc,
                "formula_ref": comparison.formula_ref,
                "formula_match": comparison.formula_match,
                "mw_cc": comparison.mw_cc,
                "mw_ref": comparison.mw_ref,
                "mw_error": comparison.mw_error,
                "logp_cc": comparison.logp_cc,
                "logp_ref": comparison.logp_ref,
                "logp_error": comparison.logp_error,
                "logp_agreement": comparison.logp_agreement,
                "tpsa_cc": comparison.tpsa_cc,
                "tpsa_ref": comparison.tpsa_ref,
                "tpsa_error": comparison.tpsa_error,
                "tpsa_agreement": comparison.tpsa_agreement,
                "rotatable_bonds_cc": comparison.rotatable_bonds_cc,
                "rotatable_bonds_ref": comparison.rotatable_bonds_ref,
                "rotatable_bonds_match": comparison.rotatable_bonds_match,
                "ring_count_cc": comparison.ring_count_cc,
                "ring_count_ref": comparison.ring_count_ref,
                "ring_count_match": comparison.ring_count_match,
            }
        )
    return rows


def summary_to_rows(summary: DescriptorBenchmarkSummary) -> list[dict[str, Any]]:
    return [
        {"metric": "Molecules benchmarked", "value": summary.n_molecules},
        {"metric": "Formula accuracy (%)", "value": f"{summary.formula_accuracy:.2f}"},
        {"metric": "Rotatable bonds agreement (%)", "value": f"{summary.rotatable_bonds_accuracy:.2f}"},
        {"metric": "Ring count agreement (%)", "value": f"{summary.ring_count_accuracy:.2f}"},
        {"metric": "Full agreement (%)", "value": f"{summary.full_agreement:.2f}"},
        {"metric": "MW MAE", "value": f"{summary.mw_mae:.4f}"},
        {"metric": "MW RMSE", "value": f"{summary.mw_rmse:.4f}"},
        {"metric": "MW max error", "value": f"{summary.mw_max_error:.4f}"},
        {"metric": "LogP MAE", "value": f"{summary.logp_mae:.4f}"},
        {"metric": "LogP RMSE", "value": f"{summary.logp_rmse:.4f}"},
        {"metric": "LogP max error", "value": f"{summary.logp_max_error:.4f}"},
        {"metric": "LogP agreement (%)", "value": f"{summary.logp_within_tolerance:.2f}"},
        {"metric": "TPSA MAE", "value": f"{summary.tpsa_mae:.4f}"},
        {"metric": "TPSA RMSE", "value": f"{summary.tpsa_rmse:.4f}"},
        {"metric": "TPSA max error", "value": f"{summary.tpsa_max_error:.4f}"},
        {"metric": "TPSA agreement (%)", "value": f"{summary.tpsa_within_tolerance:.2f}"},
    ]


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 60)


def _save_figure(fig, path: Path) -> None:
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    fig.clf()
    try:
        import matplotlib.pyplot as plt
        plt.close(fig)
    except Exception:
        pass


def build_plots(comparisons: list[DescriptorComparison], output_dir: Path, base_name: str) -> list[Path]:
    import matplotlib.pyplot as plt

    output_dir.mkdir(parents=True, exist_ok=True)
    fig_paths: list[Path] = []
    
    x_mw = [c.mw_ref for c in comparisons if c.mw_ref is not None and c.mw_cc is not None]
    y_mw = [c.mw_cc for c in comparisons if c.mw_ref is not None and c.mw_cc is not None]
    if x_mw and y_mw:
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.scatter(x_mw, y_mw, alpha=0.75, edgecolors="k", linewidths=0.4)
        min_val = min(min(x_mw), min(y_mw))
        max_val = max(max(x_mw), max(y_mw))
        ax.plot([min_val, max_val], [min_val, max_val], color="gray", linestyle="--")
        ax.set_title("MW: Chemistry Companion vs RDKit")
        ax.set_xlabel("RDKit MW")
        ax.set_ylabel("Chemistry Companion MW")
        fig_paths.append(output_dir / f"{base_name}_mw_scatter.png")
        _save_figure(fig, fig_paths[-1])

    x_logp = [c.logp_ref for c in comparisons if c.logp_ref is not None and c.logp_cc is not None]
    y_logp = [c.logp_cc for c in comparisons if c.logp_ref is not None and c.logp_cc is not None]
    if x_logp and y_logp:
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.scatter(x_logp, y_logp, alpha=0.75, edgecolors="k", linewidths=0.4)
        min_val = min(min(x_logp), min(y_logp))
        max_val = max(max(x_logp), max(y_logp))
        ax.plot([min_val, max_val], [min_val, max_val], color="gray", linestyle="--")
        ax.set_title("LogP: Chemistry Companion vs RDKit")
        ax.set_xlabel("RDKit LogP")
        ax.set_ylabel("Chemistry Companion LogP")
        fig_paths.append(output_dir / f"{base_name}_logp_scatter.png")
        _save_figure(fig, fig_paths[-1])

    x_tpsa = [c.tpsa_ref for c in comparisons if c.tpsa_ref is not None and c.tpsa_cc is not None]
    y_tpsa = [c.tpsa_cc for c in comparisons if c.tpsa_ref is not None and c.tpsa_cc is not None]
    if x_tpsa and y_tpsa:
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.scatter(x_tpsa, y_tpsa, alpha=0.75, edgecolors="k", linewidths=0.4)
        min_val = min(min(x_tpsa), min(y_tpsa))
        max_val = max(max(x_tpsa), max(y_tpsa))
        ax.plot([min_val, max_val], [min_val, max_val], color="gray", linestyle="--")
        ax.set_title("TPSA: Chemistry Companion vs RDKit")
        ax.set_xlabel("RDKit TPSA")
        ax.set_ylabel("Chemistry Companion TPSA")
        fig_paths.append(output_dir / f"{base_name}_tpsa_scatter.png")
        _save_figure(fig, fig_paths[-1])

    errors = [c.mw_error for c in comparisons if c.mw_error is not None]
    if errors:
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.hist(errors, bins=min(12, max(3, len(errors) // 2)), color="#3572A5", edgecolor="black")
        ax.set_title("MW Error Distribution")
        ax.set_xlabel("Absolute MW error")
        ax.set_ylabel("Count")
        fig_paths.append(output_dir / f"{base_name}_mw_error_hist.png")
        _save_figure(fig, fig_paths[-1])

    return fig_paths


def export_benchmark_report(
    comparisons: list[DescriptorComparison],
    output_dir: str | Path = "output",
    base_name: str = "descriptor_benchmark",
) -> dict[str, str]:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    summary = benchmark_summary(comparisons)
    summary_rows = summary_to_rows(summary)
    detail_rows = comparisons_to_rows(comparisons)

    csv_path = output_path / f"{base_name}_details.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        import csv as _csv

        fieldnames = list(detail_rows[0].keys()) if detail_rows else []
        writer = _csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(detail_rows)

    xlsx_path = output_path / f"{base_name}.xlsx"
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for row in summary_rows:
        summary_sheet.append([row["metric"], row["value"]])
    _autosize_columns(summary_sheet)

    details_sheet = wb.create_sheet(title="Details")
    if detail_rows:
        headers = list(detail_rows[0].keys())
        details_sheet.append(headers)
        for row in detail_rows:
            details_sheet.append([row.get(key, "") for key in headers])
        _autosize_columns(details_sheet)

    charts_sheet = wb.create_sheet(title="Figures")
    fig_paths = build_plots(comparisons, output_path, base_name)
    for idx, fig_path in enumerate(fig_paths, start=1):
        img = OpenpyxlImage(str(fig_path))
        cell = f"A{(idx - 1) * 25 + 1}"
        charts_sheet.add_image(img, cell)

    wb.save(xlsx_path)

    return {
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
        **{f"figure_{i+1}": str(path) for i, path in enumerate(fig_paths)},
    }


__all__ = [
    "DescriptorBenchmarkSummary",
    "DescriptorComparison",
    "benchmark_molecule",
    "benchmark_from_csv",
    "benchmark_summary",
    "export_benchmark_report",
    "compute_rdkit_reference_descriptors",
    "compare_descriptors",
]
